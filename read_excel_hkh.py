# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import json
import re

# openpyxl는 index 1 부터 시작함
START_COLUMN_INDEX = 1

not_all_none = False


def cell_value(row, column):
    global not_all_none
    val = sheet.cell(row=row, column=START_COLUMN_INDEX+column).value
    if type(val) is not str:
        val = str(val)
    val = val.replace("\"", "").rstrip()

    if val == 'None':
        val = None

    if not not_all_none:
        not_all_none = val is not None

    return val


def parse_goal_label(lienum, goal_label_str):
    items = re.findall(r"([^{:,\s]+[^{:,]*):([^,}]*)", goal_label_str)
    dic = dict()
    for item in items:
        if item[1] == 'TRUE':
            dic[item[0]] = 'True'
        elif item[1] == 'FALSE':
            dic[item[0]] = 'False'
        else:
            dic[item[0]] = item[1].strip()
    return dic


def convert2dict(session):
    utters = []
    for utter in session.utters:
        dialog_acts = []
        for dialog_act in utter.dialog_acts:
            dialog_act_dict = {
                "act": dialog_act.act,
                "slot": dialog_act.slot,
                "value": dialog_act.value,
            }
            dialog_acts += [dialog_act_dict]

        utter_dict = {
            "turn_index": utter.turn_index,
            "topic": utter.topic,
            "target_bio": utter.target_bio,
            "speaker": utter.speaker,
            "text": utter.text,
            "semantic_tagged": utter.semantic_tagged,
            "dialog_acts": dialog_acts,
            "goal_label": utter.goal_label,
            "method": utter.method,
            "requested": utter.requested,
        }
        utters += [utter_dict]

    session_dict = {
        "session_id": session.session_id,
        "utters": utters
    }
    return session_dict


# ======================================================================================================================
# ======================================================================================================================
# ======================================================================================================================
# ======================================================================================================================


# 여기 부터가 메인코드
if __name__ == '__main__':

    sheet_list = ['소개', '정보 문의', '요청', '알림', '일정 문의', '대답', '대화']
    #sheet_list = ['일정 문의']

    file = './excel/치매케어_사용자발화문_20181018_v3.xlsx'
    wb = load_workbook(filename=file)

    exception_lines = []

    sessions = []
    session = None
    init_state = True


    result_json_list = []


    # 시트 순서대로 읽어들인다.
    for sheet_name in sheet_list:

        sheet = wb[sheet_name]

        # read cells from excel
        for row in range(3, 25290):
            unit_dict = {}
            utters_dict = {}
            dialog_acts_dict = {}


            num = cell_value(row, 0)
            user_speech = cell_value(row, 1)
            intention = cell_value(row, 2)
            topic = cell_value(row, 3)
            dialog_act = cell_value(row, 4)
            slot = cell_value(row, 5)
            value = cell_value(row, 6)

            # 비어있는 행이면
            if num == None and user_speech == None and intention == None and topic == None and dialog_act == None and slot == None and value == None:
                end_of_sheet = 1
                temp = row + 1
                # 빈행으로 부터 20줄 정도가 모두 빈행인지 확인
                while temp < row + 20:
                    if cell_value(temp, 1) != None:
                        end_of_sheet = 0
                    temp += 1

                # 시트 끝까지 다 읽었으면 다음 시트로
                if end_of_sheet == 1:
                    break
                # 비어있는 행이기 때문에 다음 행으로
                continue

            # slot-value 쌍이 2개 이상 나오는 형태
            elif num == None and user_speech == None and intention == None and topic == None and dialog_act == None and slot != None and value != None:
                dialog_acts_dict['act'] = pre_dialog_act
                dialog_acts_dict['slot'] = slot
                dialog_acts_dict['value'] = value

                # 방금 전에 추가한 리스트 내 dialog_acts 리스트에 append
                result_json_list[appended_idx]['utters'][0]['dialog_acts'].append(dialog_acts_dict)
                continue


            # 이전 row의 dialog-act
            pre_dialog_act = dialog_act


            # 가장 중요하다고 생각되는 세가지 정보가 비어있는 행은 생략
            if dialog_act == None and slot == None and value == None:
                continue


            # 현재 코퍼스에 있는 slot, value 값이 온톨로지에 있는지 확인하고 없으면 추가하는 단계
            if slot != None and value != None:

                is_exist_ont_header = 0
                is_exist_ont_content = 0

                # 온톨로지에 있는지 확인하기 위해 시트 이동
                sheet = wb['Ontology']
                for col in range(0, 25290):
                    if cell_value(1, col) == None:
                        new_col_idx = col
                        break
                    if cell_value(1, col) == slot:
                        fit_col_idx = col
                        is_exist_ont_header = 1
                        for row in range(2, 25290):
                            if cell_value(row, col) == None:
                                new_row_idx = row
                                break
                            if cell_value(row, col) == value:
                                is_exist_ont_content = 1
                                break

                # slot이 없는 경우
                if is_exist_ont_header == 0:
                    sheet.cell(row=1, column=new_col_idx + START_COLUMN_INDEX).value = slot
                    sheet.cell(row=2, column=new_col_idx + START_COLUMN_INDEX).value = value
                    wb.save(file)
                    print('새로운 "' + slot + '" slot에 새로운 "' + value + '" value 추가')
                # slot은 있는데 value가 없는 경우
                elif is_exist_ont_header == 1:
                    if is_exist_ont_content == 0:
                        sheet.cell(row=new_row_idx, column=fit_col_idx + START_COLUMN_INDEX).value = value
                        wb.save(file)
                        print('기존 "' + slot + '" slot에 새로운 "' + value + '" value 추가')


                # 다시 원래 시트로 복귀
                sheet = wb[sheet_name]


            dialog_acts_dict['act'] = dialog_act
            dialog_acts_dict['slot'] = slot
            dialog_acts_dict['value'] = value


            utters_dict['dialog_acts'] = [dialog_acts_dict]
            utters_dict['semantic_tagged'] = ''
            utters_dict['speaker'] = 'User'
            utters_dict['target_bio'] = ''
            utters_dict['text'] = user_speech
            utters_dict['text_spaced'] = ''
            utters_dict['topic'] = topic
            utters_dict['turn_index'] = '1'

            unit_dict['session_id'] = ''
            unit_dict['utters'] = [utters_dict]

            result_json_list.append(unit_dict)

            # 방금 리스트에 append한 index
            appended_idx = len(result_json_list) - 1


    result_json = json.dumps(result_json_list, indent=4, ensure_ascii=False, sort_keys=True)

    f = open('./json/result_json.json', 'w', encoding='utf-8')
    f.write(result_json)
    f.close()


    # ==================================================================================================================


    # Ontology시트 데이터 json파일 생성
    ontology_dict = {}
    sheet = wb['Ontology']
    for col in range(0, 25290):

        temp_list = []

        if cell_value(1, col) == None:
            break

        for row in range(2, 25290):
            if cell_value(row, col) == None:
                break

            # 공백을 포함한 ontology는 공백 제거한 것도 ontology 리스트에 추가시킨다.
            if ' ' in cell_value(row, col):
                temp_cell_value = cell_value(row, col).replace(' ', '')
                if temp_cell_value not in temp_list:
                    temp_list.append(temp_cell_value)

            temp_list.append(cell_value(row, col))

        ontology_dict[cell_value(1,col)] = temp_list

    ontology_json = json.dumps(ontology_dict, indent=4, ensure_ascii=False, sort_keys=True)

    f = open('./json/ontology_json.json', 'w', encoding='utf-8')
    f.write(ontology_json)
    f.close()



